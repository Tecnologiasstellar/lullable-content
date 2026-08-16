#!/usr/bin/env python3
"""
lullable.py — one tool for the Lullable story pipeline.

CANONICAL SOURCE OF TRUTH: Stories/<storyID>/story.yaml
Everything else is generated FROM it and must never be hand-edited:
    _generated/tracker-row.tsv     one line for the Excel tracker
    _generated/story-card.yaml     the app-facing card
    _generated/catalog-payload.json  the Supabase upsert body
    _generated/neon-metadata.json    the Neon record
    Lullable_Story_Card_Tracker.xlsx rebuilt from all manifests

Subcommands
    compile   narration.md -> SSML + creates/updates story.yaml
    validate  run the staged gate model over one story or all
    build     regenerate every derived artifact from story.yaml
    tracker   rebuild the Excel workbook from all manifests
"""
import argparse, hashlib, json, os, random, re, subprocess, sys, datetime
import yaml

SCHEMA_VERSION = 1

# ---------------------------------------------------------------- compiler ---
INTRA = [(1.0,.28),(1.2,.30),(1.5,.22),(1.8,.14),(2.0,.06)]
PARA  = [(2.0,.30),(2.2,.22),(2.4,.28),(2.6,.12),(3.0,.08)]
SEAM  = 3.0
BREATH_MIN, BREATH_MAX = 5, 9
MAX_WORDS_PER_BREAK = 15.0
MIN_RUNTIME_MIN     = 45.0
MAX_RUNTIME_MIN     = 65.0
LONG_SENTENCE = 22

# ------------------------------------------------------------- vocabularies --
GENRES            = ("ancient-worlds","gentle-nature","cosmic-journeys","cozy-tales")
WORKFLOW_STATUSES = ("draft","rendered","qa-approved","staging","published")
ACCESS_DECISIONS  = ("PENDING","free","premium")
RIGHTS_STATUSES   = ("pending-verification","verified","restricted")
V2_MODELS         = ("eleven_multilingual_v2","eleven_english_v2","eleven_turbo_v2","eleven_turbo_v2_5")
PLACEHOLDERS      = ("PENDING","VOICE_ID","TODO","XXX","CHANGEME","")

DELIVERY_SPEC = {"codec":"aac","profile":"LC","sampleRate":44100,"channels":1,"bitRateKbps":96}
BITRATE_MIN_KBPS, BITRATE_MAX_KBPS = 58, 140   # 96 nominal; VBR on speech drifts, so this
                                               # band catches real mistakes (32k, 320k stereo)
                                               # without failing normal encoder variation
DURATION_TOLERANCE_S   = 1.0

CARD_TARGETS = {   # field: (editorial min, editorial max, hard cap)
    "title":(18,48,200), "subtitle":(35,90,300), "narrator":(2,32,160),
    "bedtimeNote":(90,180,320), "bestFor":(10,24,28), "sleepPace":(8,22,24),
    "atmosphere":(10,28,30), "description":(220,650,4000)}

# tracker columns, in order. This list is the ONE definition of the row shape.
TRACKER_COLUMNS = [
    "storyID","supersedes","workflowStatus","accessDecision","access","title","subtitle",
    "narrator","genreID_1","genreID_2","trialPreviewEligible","isFeatured","publishedAt",
    "colorHex","accentHex","durationSeconds","bedtimeNote","bestFor","sleepPace","atmosphere",
    "description","audioAssetID","audioMasterFilename","commercialRightsStatus","audioDelivery"]

def _p(*a): print(*a)

# ----------------------------------------------------------------- helpers ---
def wc(s): return len(re.findall(r"[A-Za-z]+", s or ""))
def split_sentences(par): return [p.strip() for p in re.split(r'(?<=[.?…])\s+', par.strip()) if p.strip()]
def slugify(t): return re.sub(r"[^A-Za-z0-9]+","_",t).strip("_").upper() or "EPISODE"
def story_id_from(t): return re.sub(r"-+","-",re.sub(r"[^a-z0-9]+","-",(t or "").lower())).strip("-")[:80]
def is_placeholder(v): return (str(v).strip() in PLACEHOLDERS) or str(v).strip().upper().startswith("PENDING")
def weighted(ch, rng):
    v,w = zip(*ch); return rng.choices(v, weights=w, k=1)[0]

def sha256_of(path, chunk=1 << 20):
    h = hashlib.sha256()
    with open(path,"rb") as f:
        for b in iter(lambda: f.read(chunk), b""): h.update(b)
    return h.hexdigest()

def ffprobe(path):
    """Return a dict of real encoding facts, or None if unreadable."""
    try:
        out = subprocess.run(
            ["ffprobe","-v","error","-print_format","json","-show_format","-show_streams",path],
            capture_output=True, text=True, timeout=60)
        if out.returncode != 0: return None
        d = json.loads(out.stdout)
        a = next((s for s in d.get("streams",[]) if s.get("codec_type")=="audio"), None)
        if not a: return None
        fmt = d.get("format",{})
        br = a.get("bit_rate") or fmt.get("bit_rate")
        return {"codec": a.get("codec_name"),
                "profile": (a.get("profile") or "").replace("LC","LC"),
                "sampleRate": int(a.get("sample_rate") or 0),
                "channels": int(a.get("channels") or 0),
                "durationSeconds": round(float(fmt.get("duration") or a.get("duration") or 0), 3),
                "bitRateKbps": round(int(br)/1000) if br else None,
                "bytes": int(fmt.get("size") or 0)}
    except Exception:
        return None

def parse_iso_utc(s):
    """Strict ISO-8601 UTC, e.g. 2026-09-15T16:00:00Z."""
    if not isinstance(s,str) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z", s):
        return None
    try: return datetime.datetime.strptime(s, "%Y-%m-%dT%H:%M:%SZ")
    except ValueError: return None

# ------------------------------------------------------------------- YAML ----
def dump_yaml(obj):
    """Safe emission. PyYAML handles quotes, colons, newlines and unicode."""
    return yaml.safe_dump(obj, sort_keys=False, allow_unicode=True, width=100000,
                          default_flow_style=False)

def load_manifest(story_dir):
    p = os.path.join(story_dir, "story.yaml")
    if not os.path.exists(p): raise FileNotFoundError(p)
    with open(p, encoding="utf-8") as f: return yaml.safe_load(f)

def save_manifest(story_dir, m):
    with open(os.path.join(story_dir,"story.yaml"),"w",encoding="utf-8") as f:
        f.write("# CANONICAL MANIFEST — edit this file and nothing else.\n")
        f.write("# Everything in _generated/ is produced from here by: lullable.py build\n\n")
        f.write(dump_yaml(m))

def blank_manifest():
    return {
      "schemaVersion": SCHEMA_VERSION,
      "storyID": "PENDING", "supersedes": None, "identityResolved": False,
      "identityNote": "",
      "workflowStatus": "draft", "accessDecision": "PENDING",
      "episode": {"id": 0, "pillar": "", "tags": []},
      "card": {"title":"PENDING","subtitle":"PENDING","narrator":"PENDING",
               "genreIDs":[], "bedtimeNote":"PENDING","bestFor":"PENDING",
               "sleepPace":"PENDING","atmosphere":"PENDING","description":"PENDING",
               "colorHex":"PENDING","accentHex":"PENDING",
               "isFeatured": False, "trialPreviewEligible": False,
               "publishedAt":"PENDING", "durationSeconds": None},
      "script": {"narrationFile":"narration.md","ssmlFile":"upload-to-elevenlabs.txt",
                 "words":0,"breaks":0,"wordsPerBreak":0.0,"silenceSeconds":0.0,
                 "estimatedMinutesAt118wpm":0.0},
      "rights": {"status":"verified",
                 "evidence":("Script written in-house for Lullable; narration rendered with our own "
                             "licensed ElevenLabs voice; no third-party text, music or sound used. "
                             "All rights owned by Lullable."),
                 "evidenceFiles": []},
      "render": {"provider":"elevenlabs","voiceId":"PENDING","voiceName":"PENDING",
                 "model":"PENDING",
                 "settings":{"stability":None,"similarityBoost":None,"style":None,"speakerBoost":None},
                 "historyItemId":"PENDING","projectId":"PENDING","renderedAt":"PENDING"},
      "audio": {"master":{"filename":"PENDING","sha256":"PENDING","bytes":None,
                          "durationSeconds":None,"codec":None,"sampleRate":None,"channels":None},
                "delivery":{"filename":"PENDING","sha256":"PENDING","bytes":None,
                            "durationSeconds":None,"codec":None,"profile":None,
                            "sampleRate":None,"channels":None,"bitRateKbps":None}},
      "qa": {"audioApproved": False,"approvedBy":"","approvedAt":"",
             "deviceAccepted": False,"deviceNotes":""},
      "publish": {"audioAssetID":"PENDING","supabaseAudioUploaded": False,
                  "catalogRowUpserted": False},
    }

# ================================================================== GATES ====
# Each gate returns (status, message). status: PASS | FAIL | NA
# STAGE_GATES says which gates a story must pass to legitimately be at a stage.

def _get(m, path, default=None):
    cur = m
    for k in path.split("."):
        if not isinstance(cur, dict) or k not in cur: return default
        cur = cur[k]
    return cur

def g01_schema(m, d):
    missing = [k for k in ("schemaVersion","storyID","workflowStatus","accessDecision",
                           "card","script","rights","render","audio","qa","publish") if k not in m]
    if missing: return "FAIL", "missing top-level keys: " + ", ".join(missing)
    if m.get("schemaVersion") != SCHEMA_VERSION:
        return "FAIL", "schemaVersion %r, expected %d" % (m.get("schemaVersion"), SCHEMA_VERSION)
    return "PASS", "schema v%d, all sections present" % SCHEMA_VERSION

def g02_identity(m, d):
    sid = m.get("storyID","")
    if is_placeholder(sid): return "FAIL", "storyID is unset"
    if not re.fullmatch(r"[a-z0-9-]{3,80}", sid):
        return "FAIL", "storyID %r is not 3-80 chars of lowercase/digits/hyphen" % sid
    if not m.get("identityResolved"):
        note = m.get("identityNote") or "identityResolved is false"
        return "FAIL", "story identity unresolved — " + note
    sup = m.get("supersedes")
    if sup and not re.fullmatch(r"[a-z0-9-]{3,80}", sup):
        return "FAIL", "supersedes %r is not a valid storyID" % sup
    if sup == sid: return "FAIL", "supersedes equals storyID"
    return "PASS", sid + ((" (supersedes %s)" % sup) if sup else "")

def g03_enums(m, d):
    bad = []
    if m.get("workflowStatus") not in WORKFLOW_STATUSES:
        bad.append("workflowStatus=%r" % m.get("workflowStatus"))
    if m.get("accessDecision") not in ACCESS_DECISIONS:
        bad.append("accessDecision=%r" % m.get("accessDecision"))
    if _get(m,"rights.status") not in RIGHTS_STATUSES:
        bad.append("rights.status=%r" % _get(m,"rights.status"))
    gids = _get(m,"card.genreIDs") or []
    if not isinstance(gids, list) or not (1 <= len(gids) <= 2):
        bad.append("card.genreIDs must be a list of 1-2 ids")
    else:
        for g in gids:
            if g not in GENRES: bad.append("genreID=%r" % g)
    for b in ("card.isFeatured","card.trialPreviewEligible","qa.audioApproved",
              "qa.deviceAccepted","publish.supabaseAudioUploaded","publish.catalogRowUpserted"):
        if not isinstance(_get(m,b), bool): bad.append("%s must be a real boolean" % b)
    return ("FAIL", "; ".join(bad)) if bad else ("PASS","all enums and booleans valid")

def g04_copy(m, d):
    """Authored copy only. `narrator` is set by voice assignment at render time,
    so a PENDING narrator is legitimate before that; G07/G12 catch it later."""
    bad = []
    for f,(lo,hi,cap) in CARD_TARGETS.items():
        v = _get(m,"card."+f, "")
        if is_placeholder(v):
            if f == "narrator" and m.get("workflowStatus") == "draft": continue
            bad.append("%s is PENDING" % f); continue
        n = len(v)
        if n > cap: bad.append("%s %d chars OVER HARD CAP %d" % (f,n,cap))
        elif not (lo <= n <= hi): bad.append("%s %d chars, target %d-%d" % (f,n,lo,hi))
    return ("FAIL","; ".join(bad)) if bad else ("PASS","all 8 copy fields within editorial targets")

def g05_hex(m, d):
    bad = [k for k in ("colorHex","accentHex")
           if not re.fullmatch(r"[0-9A-F]{6}", str(_get(m,"card."+k,"")))]
    return ("FAIL","not 6 uppercase hex chars: " + ", ".join(bad)) if bad else ("PASS","colour values valid")

def g06_publishedat(m, d):
    v = _get(m,"card.publishedAt")
    if is_placeholder(v): return "FAIL","publishedAt is PENDING"
    if parse_iso_utc(v) is None:
        return "FAIL","publishedAt %r is not strict ISO-8601 UTC (YYYY-MM-DDTHH:MM:SSZ)" % v
    return "PASS", v

def g07_no_placeholders(m, d):
    found = []
    def walk(o, path):
        if isinstance(o, dict):
            for k,v in o.items(): walk(v, path+"."+k if path else k)
        elif isinstance(o, list):
            for i,v in enumerate(o): walk(v, "%s[%d]" % (path,i))
        elif isinstance(o,str) and o.strip() and is_placeholder(o):
            found.append(path)
    walk(m, "")
    return ("FAIL","placeholder values remain: " + ", ".join(sorted(found))) if found else ("PASS","no placeholders anywhere")

def _audio_path(d, m, which):
    fn = _get(m,"audio.%s.filename" % which)
    if is_placeholder(fn): return None
    return os.path.join(d, "audio", fn)

def g08_files(m, d):
    msgs, bad = [], []
    for which in ("master","delivery"):
        p = _audio_path(d,m,which)
        if p is None: bad.append("%s filename unset" % which); continue
        if not os.path.exists(p): bad.append("%s missing on disk: audio/%s" % (which, os.path.basename(p)))
        else: msgs.append("%s present (%.1f MB)" % (which, os.path.getsize(p)/1e6))
    for ev in (_get(m,"rights.evidenceFiles") or []):
        if not os.path.exists(os.path.join(d,"rights",ev)):
            bad.append("rights evidence missing: rights/%s" % ev)
    return ("FAIL","; ".join(bad)) if bad else ("PASS","; ".join(msgs) or "no files required")

def g09_hashes(m, d):
    bad, ok = [], []
    for which in ("master","delivery"):
        p = _audio_path(d,m,which)
        rec = _get(m,"audio.%s.sha256" % which)
        if p is None or not os.path.exists(p): bad.append("%s not on disk" % which); continue
        if is_placeholder(rec): bad.append("%s sha256 not recorded" % which); continue
        actual = sha256_of(p)
        if actual != rec: bad.append("%s CHECKSUM MISMATCH (file has %s…)" % (which, actual[:12]))
        else: ok.append("%s %s…" % (which, actual[:12]))
    return ("FAIL","; ".join(bad)) if bad else ("PASS","; ".join(ok))

def g10_encoding(m, d):
    p = _audio_path(d,m,"delivery")
    if p is None or not os.path.exists(p): return "FAIL","delivery file not available to probe"
    probe = ffprobe(p)
    if not probe: return "FAIL","ffprobe could not read the delivery file"
    bad = []
    if probe["codec"] != DELIVERY_SPEC["codec"]: bad.append("codec %s, want aac" % probe["codec"])
    if DELIVERY_SPEC["profile"] not in (probe["profile"] or ""): bad.append("profile %r, want LC" % probe["profile"])
    if probe["sampleRate"] != DELIVERY_SPEC["sampleRate"]: bad.append("%d Hz, want 44100" % probe["sampleRate"])
    if probe["channels"] != DELIVERY_SPEC["channels"]: bad.append("%d channels, want mono" % probe["channels"])
    br = probe["bitRateKbps"]
    if br is None or not (BITRATE_MIN_KBPS <= br <= BITRATE_MAX_KBPS):
        bad.append("%s kbps, outside the %d-%d band (nominal 96)" % (br, BITRATE_MIN_KBPS, BITRATE_MAX_KBPS))
    mp = _audio_path(d,m,"master")
    if mp and os.path.exists(mp):
        mpr = ffprobe(mp)
        if mpr and not str(mpr["codec"]).startswith("pcm"):
            bad.append("master codec %s, want uncompressed PCM/WAV" % mpr["codec"])
    return ("FAIL","; ".join(bad)) if bad else \
           ("PASS","AAC-LC %d Hz mono %s kbps" % (probe["sampleRate"], br))

def g11_duration(m, d):
    card = _get(m,"card.durationSeconds")
    p = _audio_path(d,m,"delivery")
    if not isinstance(card,(int,float)): return "FAIL","card.durationSeconds is not a number"
    if p is None or not os.path.exists(p): return "FAIL","no delivery file to measure against"
    probe = ffprobe(p)
    if not probe: return "FAIL","could not measure delivery duration"
    diff = abs(probe["durationSeconds"] - card)
    if diff > DURATION_TOLERANCE_S:
        return "FAIL","card says %ss, audio is %.1fs (%.1fs apart)" % (card, probe["durationSeconds"], diff)
    return "PASS","%ss, matches audio within %.1fs" % (card, diff)

def g12_render(m, d):
    bad = []
    for k in ("voiceId","voiceName","model","renderedAt"):
        v = _get(m,"render."+k)
        if is_placeholder(v): bad.append("render.%s unset" % k)
    history_id = _get(m,"render.historyItemId")
    project_id = _get(m,"render.projectId")
    chapter_id = _get(m,"render.chapterId")
    present = lambda value: value is not None and bool(str(value).strip()) and not is_placeholder(value)
    has_history_evidence = present(history_id)
    has_studio_evidence = present(project_id) and present(chapter_id)
    if not (has_history_evidence or has_studio_evidence):
        bad.append("render provenance requires historyItemId or both projectId and chapterId")
    model = _get(m,"render.model")
    if not is_placeholder(model) and model not in V2_MODELS:
        bad.append("model %r is not a v2-family model that honours SSML breaks" % model)
    st = _get(m,"render.settings") or {}
    if any(st.get(k) is None for k in ("stability","similarityBoost")):
        bad.append("render.settings stability/similarityBoost not recorded")
    ra = _get(m,"render.renderedAt")
    if not is_placeholder(ra) and parse_iso_utc(ra) is None:
        bad.append("render.renderedAt is not ISO-8601 UTC")
    provenance = "history item" if has_history_evidence else "Studio project + chapter"
    return ("FAIL","; ".join(bad)) if bad else ("PASS","%s on %s; %s recorded" % (_get(m,"render.voiceName"), model, provenance))

def g13_qa(m, d):
    bad = []
    if not _get(m,"qa.audioApproved"): bad.append("audio QA not approved")
    elif is_placeholder(_get(m,"qa.approvedBy")): bad.append("qa.approvedBy not recorded")
    elif parse_iso_utc(_get(m,"qa.approvedAt") or "") is None: bad.append("qa.approvedAt not ISO-8601 UTC")
    if not _get(m,"qa.deviceAccepted"): bad.append("not accepted on a physical device")
    return ("FAIL","; ".join(bad)) if bad else ("PASS","approved by %s; device accepted" % _get(m,"qa.approvedBy"))

def g14_publish(m, d):
    bad = []
    if is_placeholder(_get(m,"publish.audioAssetID")): bad.append("audioAssetID not minted")
    if not _get(m,"publish.supabaseAudioUploaded"): bad.append("audio not uploaded to Supabase")
    if not _get(m,"publish.catalogRowUpserted"): bad.append("catalog row not upserted")
    return ("FAIL","; ".join(bad)) if bad else ("PASS","asset %s live in catalog" % _get(m,"publish.audioAssetID"))

def g15_rights(m, d):
    st = _get(m,"rights.status")
    if st != "verified": return "FAIL","rights status is %r, must be verified to publish" % st
    if is_placeholder(_get(m,"rights.evidence")): return "FAIL","rights.evidence is empty"
    return "PASS","verified; evidence recorded"

def g16_access(m, d):
    dec = m.get("accessDecision")
    if dec == "PENDING":
        return "FAIL","accessDecision still PENDING — free vs premium must be settled before staging"
    return "PASS","access will publish as %r" % dec

def g17_ssml(m, d):
    p = os.path.join(d, _get(m,"script.ssmlFile") or "upload-to-elevenlabs.txt")
    if not os.path.exists(p): return "FAIL","SSML file missing: %s" % os.path.basename(p)
    s = open(p, encoding="utf-8").read()
    bad = []
    tags = set(re.findall(r'<([^ >/]+)', s))
    if tags - {"break"}: bad.append("non-break tags present: %s" % ", ".join(sorted(tags-{"break"})))
    if not s.startswith("Good evening."): bad.append("does not open on 'Good evening.'")
    if not s.rstrip().endswith('Goodnight. <break time="3.0s"/>'): bad.append("does not close on 'Goodnight.'")
    if "§" in s: bad.append("section marker leaked into the SSML")
    if "!" in s: bad.append("exclamation mark present")
    wpb = _get(m,"script.wordsPerBreak") or 0
    if wpb > MAX_WORDS_PER_BREAK: bad.append("words/break %.1f above %.1f" % (wpb, MAX_WORDS_PER_BREAK))
    return ("FAIL","; ".join(bad)) if bad else ("PASS","break-only, %.1f words/break" % wpb)

GATES = [
 ("G01","manifest schema",        g01_schema),
 ("G02","story identity",         g02_identity),
 ("G03","allowed values",         g03_enums),
 ("G04","card copy lengths",      g04_copy),
 ("G05","artwork colours",        g05_hex),
 ("G06","publish date ISO-8601",  g06_publishedat),
 ("G07","no placeholders left",   g07_no_placeholders),
 ("G08","audio + rights files",   g08_files),
 ("G09","audio checksums",        g09_hashes),
 ("G10","delivery encoding",      g10_encoding),
 ("G11","duration matches audio", g11_duration),
 ("G12","render manifest",        g12_render),
 ("G13","QA + device sign-off",   g13_qa),
 ("G14","Supabase + catalog",     g14_publish),
 ("G15","commercial rights",      g15_rights),
 ("G16","access decision final",  g16_access),
 ("G17","SSML integrity",         g17_ssml),
]

STAGE_GATES = {
 "draft":       ["G01","G02","G03","G04","G05","G15","G17"],
 "rendered":    ["G01","G02","G03","G04","G05","G15","G17","G08","G09","G10","G11","G12"],
 "qa-approved": ["G01","G02","G03","G04","G05","G15","G17","G08","G09","G10","G11","G12","G13"],
 "staging":     ["G01","G02","G03","G04","G05","G15","G17","G08","G09","G10","G11","G12","G13","G16"],
 "published":   [g[0] for g in GATES],
}

def run_gates(m, d):
    results = {}
    for gid, title, fn in GATES:
        try: status, msg = fn(m, d)
        except Exception as e: status, msg = "FAIL", "gate error: %s" % e
        results[gid] = (title, status, msg)
    return results

def stage_verdict(m, results, stage=None):
    stage = stage or m.get("workflowStatus","draft")
    required = STAGE_GATES.get(stage, STAGE_GATES["draft"])
    failed = [g for g in required if results[g][1] != "PASS"]
    return stage, required, failed

def publish_ready(results):
    return all(results[g][1] == "PASS" for g in STAGE_GATES["published"])

# ============================================================== GENERATORS ===
# Every function below derives from the manifest. Nothing here is authored.

def derived_access(m):
    """`access` is only final at staging/published; before that it is PENDING."""
    dec = m.get("accessDecision","PENDING")
    if dec == "PENDING": return "PENDING"
    if m.get("workflowStatus") in ("staging","published"): return dec
    return "PENDING"

def tracker_row(m):
    g = (_get(m,"card.genreIDs") or []) + ["",""]
    dur = _get(m,"card.durationSeconds")
    return {
      "storyID": m.get("storyID",""),
      "supersedes": m.get("supersedes") or "",
      "workflowStatus": m.get("workflowStatus",""),
      "accessDecision": m.get("accessDecision",""),
      "access": derived_access(m),
      "title": _get(m,"card.title",""),
      "subtitle": _get(m,"card.subtitle",""),
      "narrator": _get(m,"card.narrator",""),
      "genreID_1": g[0], "genreID_2": g[1],
      "trialPreviewEligible": str(bool(_get(m,"card.trialPreviewEligible"))).lower(),
      "isFeatured": str(bool(_get(m,"card.isFeatured"))).lower(),
      "publishedAt": _get(m,"card.publishedAt",""),
      "colorHex": _get(m,"card.colorHex",""), "accentHex": _get(m,"card.accentHex",""),
      "durationSeconds": dur if isinstance(dur,(int,float)) else "PENDING",
      "bedtimeNote": _get(m,"card.bedtimeNote",""), "bestFor": _get(m,"card.bestFor",""),
      "sleepPace": _get(m,"card.sleepPace",""), "atmosphere": _get(m,"card.atmosphere",""),
      "description": _get(m,"card.description",""),
      "audioAssetID": _get(m,"publish.audioAssetID",""),
      "audioMasterFilename": _get(m,"audio.master.filename",""),
      "commercialRightsStatus": _get(m,"rights.status",""),
      "audioDelivery": "AAC-LC M4A, 44.1 kHz, mono, 96 kbps",
    }

def story_card(m):
    """The app-facing card, in the v1 template's field order. Safely emitted."""
    dur = _get(m,"card.durationSeconds")
    return {
      "storyID": m.get("storyID"), "title": _get(m,"card.title"),
      "subtitle": _get(m,"card.subtitle"), "narrator": _get(m,"card.narrator"),
      "genreIDs": list(_get(m,"card.genreIDs") or []),
      "access": derived_access(m),
      "trialPreviewEligible": bool(_get(m,"card.trialPreviewEligible")),
      "isFeatured": bool(_get(m,"card.isFeatured")),
      "publishedAt": _get(m,"card.publishedAt"),
      "colorHex": _get(m,"card.colorHex"), "accentHex": _get(m,"card.accentHex"),
      "durationSeconds": dur if isinstance(dur,(int,float)) else "PENDING",
      "bedtimeNote": _get(m,"card.bedtimeNote"), "bestFor": _get(m,"card.bestFor"),
      "sleepPace": _get(m,"card.sleepPace"), "atmosphere": _get(m,"card.atmosphere"),
      "description": _get(m,"card.description"),
      "audioAssetID": _get(m,"publish.audioAssetID"),
      "audioMasterFilename": _get(m,"audio.master.filename"),
      "commercialRightsStatus": _get(m,"rights.status"),
      "rightsEvidence": _get(m,"rights.evidence"),
      "audioDelivery": "AAC-LC M4A, 44.1 kHz, mono, 96 kbps",
    }

def catalog_payload(m, results):
    """Supabase upsert body. Refuses to build unless the story is fit to publish."""
    stage, required, failed = stage_verdict(m, results, "staging")
    if failed:
        return {"_refused": True,
                "_reason": "not fit for catalog: " + ", ".join("%s %s" % (f, results[f][0]) for f in failed)}
    c = story_card(m)
    return {"story_id": c["storyID"], "supersedes": m.get("supersedes"),
            "title": c["title"], "subtitle": c["subtitle"], "narrator": c["narrator"],
            "genre_ids": c["genreIDs"], "access": c["access"],
            "trial_preview_eligible": c["trialPreviewEligible"], "is_featured": c["isFeatured"],
            "published_at": c["publishedAt"], "color_hex": c["colorHex"], "accent_hex": c["accentHex"],
            "duration_seconds": c["durationSeconds"], "bedtime_note": c["bedtimeNote"],
            "best_for": c["bestFor"], "sleep_pace": c["sleepPace"], "atmosphere": c["atmosphere"],
            "description": c["description"], "audio_asset_id": c["audioAssetID"],
            "audio_sha256": _get(m,"audio.delivery.sha256")}

def neon_metadata(m, story_dir):
    ssml = ""
    p = os.path.join(story_dir, _get(m,"script.ssmlFile") or "")
    if os.path.exists(p): ssml = open(p, encoding="utf-8").read()
    return {"id": _get(m,"episode.id"), "storyID": m.get("storyID"),
            "title": _get(m,"card.title"), "pillar": _get(m,"episode.pillar"),
            "description": _get(m,"card.description"), "tags": _get(m,"episode.tags") or [],
            "workflowStatus": m.get("workflowStatus"),
            "estimated_reading_time": "~%.0f min at 118 wpm (incl. %.1f min breaks)" % (
                _get(m,"script.estimatedMinutesAt118wpm") or 0, (_get(m,"script.silenceSeconds") or 0)/60),
            "script_body": ssml,
            "render": m.get("render"), "audio": m.get("audio"),
            "tts_model_note": "Upload via ElevenLabs Studio; use a v2-family model that honors SSML break tags (NOT Eleven v3)."}

def publish_commands(m, results):
    """The commands to actually ship. Generated from the manifest, never typed."""
    stage, req, failed = stage_verdict(m, results, "staging")
    sid = m.get("storyID"); dv = _get(m,"audio.delivery.filename")
    if failed:
        return ["# BLOCKED — %d gate(s) failing for staging:" % len(failed)] + \
               ["#   %s %s — %s" % (f, results[f][0], results[f][2]) for f in failed]
    return [
      "# 1. upload the delivery file",
      "supabase storage cp audio/%s ss:///story-audio/%s/%s" % (dv, sid, dv),
      "# 2. upsert the catalog row",
      "psql \"$SUPABASE_DB_URL\" -c \"\\copy stories from program 'cat _generated/catalog-payload.json'\"",
      "# 3. mark it published in the manifest, then rebuild",
      "lullable.py build %s" % sid,
    ]

def build_story(story_dir, quiet=False):
    m = load_manifest(story_dir)
    results = run_gates(m, story_dir)
    out = os.path.join(story_dir, "_generated"); os.makedirs(out, exist_ok=True)
    hdr = "# GENERATED FROM story.yaml — do not edit. Regenerate with: lullable.py build\n"
    with open(os.path.join(out,"story-card.yaml"),"w",encoding="utf-8") as f:
        f.write(hdr + "\n" + dump_yaml(story_card(m)))
    row = tracker_row(m)
    with open(os.path.join(out,"tracker-row.tsv"),"w",encoding="utf-8") as f:
        f.write("\t".join(str(row[c]).replace("\t"," ").replace("\n"," ") for c in TRACKER_COLUMNS) + "\n")
    with open(os.path.join(out,"catalog-payload.json"),"w",encoding="utf-8") as f:
        json.dump(catalog_payload(m, results), f, ensure_ascii=False, indent=2)
    with open(os.path.join(out,"neon-metadata.json"),"w",encoding="utf-8") as f:
        json.dump(neon_metadata(m, story_dir), f, ensure_ascii=False, indent=2)
    with open(os.path.join(out,"publish-commands.sh"),"w",encoding="utf-8") as f:
        f.write("#!/usr/bin/env bash\n# GENERATED FROM story.yaml — do not edit.\nset -euo pipefail\n\n")
        f.write("\n".join(publish_commands(m, results)) + "\n")
    if not quiet: _p("  built _generated/ for %s (5 artifacts)" % m.get("storyID"))
    return m, results

# ================================================================ COMPILER ===
def compile_body(text, seed=11):
    rng = random.Random(seed)
    sections = [s.strip() for s in re.split(r'\n\s*§[^\n]*\n', text.strip()) if s.strip()]
    out = []
    for sec in sections:
        paras = [p.strip() for p in re.split(r'\n\s*\n', sec) if p.strip()]
        for pi, par in enumerate(paras):
            sents = split_sentences(par)
            groups, cur, cwc, tgt = [], [], 0, rng.randint(BREATH_MIN, BREATH_MAX)
            for s in sents:
                cur.append(s); cwc += wc(s)
                if cwc >= tgt:
                    groups.append(" ".join(cur)); cur, cwc = [], 0
                    tgt = rng.randint(BREATH_MIN, BREATH_MAX)
            if cur: groups.append(" ".join(cur))
            last, pieces = (pi == len(paras)-1), []
            for gi, g in enumerate(groups):
                pieces.append(g)
                dur = (SEAM if last else weighted(PARA,rng)) if gi == len(groups)-1 else weighted(INTRA,rng)
                pieces.append('<break time="%.1fs"/>' % dur)
            out.append(" ".join(pieces))
    return re.sub(r' *\n\n *','\n\n', re.sub(r'[ \t]+',' ', "\n\n".join(out))).strip()

def long_sentences(text):
    plain = re.sub(r'\n\s*§[^\n]*\n','\n\n',text); out=[]
    for par in re.split(r'\n\s*\n', plain):
        for s in split_sentences(par):
            n = wc(s)
            if n >= LONG_SENTENCE: out.append((n,s))
    return sorted(out, key=lambda x:-x[0])

def cmd_compile(a):
    d = a.story_dir
    text = open(os.path.join(d, a.narration), encoding="utf-8").read()
    body = compile_body(text, seed=a.seed)
    words = len(re.findall(r"[A-Za-z]+", re.sub(r'<[^>]+>','',body)))
    brk = [float(x) for x in re.findall(r'<break time="([\d.]+)s"/>', body)]
    wpb = words/max(len(brk),1); silence = sum(brk)
    est = words*(1.0/118.0) + silence/60.0
    _p("[compile] words=%d breaks=%d words/break=%.1f silence=%.1fmin est=~%.0fmin@118wpm"
       % (words, len(brk), wpb, silence/60, est))
    p = os.path.join(d,"story.yaml")
    m = load_manifest(d) if os.path.exists(p) else blank_manifest()
    # A settled, deliberate exception to the 45-65 band, recorded in the manifest
    # rather than in someone's memory — so a rebuild cannot silently undo it.
    exempt = bool(_get(m, "episode.runtimeExempt"))
    off_runtime = not (MIN_RUNTIME_MIN <= est <= MAX_RUNTIME_MIN) and not exempt
    if exempt: _p("  note: runtime exempt — %s" % (_get(m,"episode.runtimeExemptNote") or "settled exception"))
    if (wpb > MAX_WORDS_PER_BREAK or "!" in body or off_runtime) and not a.force:
        _p("\n  QUALITY GATE FAILED — nothing written.")
        if wpb > MAX_WORDS_PER_BREAK: _p("    * words/break %.1f above %.1f" % (wpb, MAX_WORDS_PER_BREAK))
        if "!" in body: _p("    * exclamation marks present")
        if off_runtime:
            _p("    * runtime ~%.1f min outside %.0f-%.0f min (add or cut whole sections, not sentence length)"
               % (est, MIN_RUNTIME_MIN, MAX_RUNTIME_MIN))
        for n,s in long_sentences(text)[:12]:
            _p("    [%2d] %s" % (n, s[:96] + ("..." if len(s)>96 else "")))
        _p("  Split these sentences and run again. (--force overrides.)\n")
        sys.exit(1)
    open(os.path.join(d,"upload-to-elevenlabs.txt"),"w",encoding="utf-8").write(body+"\n")
    readable = re.sub(r'\n{3,}','\n\n', re.sub(r'\n\s*§[^\n]*\n','\n\n',text)).strip()
    open(os.path.join(d,"script.md"),"w",encoding="utf-8").write(readable+"\n")
    m["script"] = {"narrationFile": a.narration, "ssmlFile":"upload-to-elevenlabs.txt",
                   "words": words, "breaks": len(brk), "wordsPerBreak": round(wpb,2),
                   "silenceSeconds": round(silence,1), "estimatedMinutesAt118wpm": round(est,1)}
    save_manifest(d, m)
    _p("  gate: PASS — wrote SSML, script.md, and updated story.yaml")

# ================================================================ COMMANDS ===
def story_dirs(root):
    s = os.path.join(root,"Stories")
    return sorted(os.path.join(s,x) for x in os.listdir(s)
                  if os.path.isdir(os.path.join(s,x)) and os.path.exists(os.path.join(s,x,"story.yaml")))


def stale_artifacts(story_dir):
    """True when story.yaml is newer than what was generated from it."""
    man = os.path.join(story_dir, "story.yaml")
    gen = os.path.join(story_dir, "_generated", "tracker-row.tsv")
    if not os.path.exists(gen): return True
    return os.path.getmtime(man) > os.path.getmtime(gen) + 1

def cmd_validate(a):
    dirs = story_dirs(a.root) if a.all else [os.path.join(a.root,"Stories",a.story)]
    worst = 0
    for d in dirs:
        m = load_manifest(d); results = run_gates(m,d)
        stage, required, failed = stage_verdict(m, results)
        ready = publish_ready(results)
        _p("\n%s  [%s]  %s" % (m.get("storyID"), stage,
                               "PUBLISH READY" if ready else "NOT READY"))
        if stale_artifacts(d):
            _p("  !! _generated/ is older than story.yaml — run  lullable.py build %s" % os.path.basename(d))
            worst = 1
        _p("-"*78)
        for gid,title,fn in GATES:
            t,st,msg = results[gid]
            req = "req" if gid in required else "   "
            mark = {"PASS":"PASS","FAIL":"FAIL","NA":"n/a"}[st]
            _p("  %s %s %-4s %-24s %s" % (gid, req, mark, t, msg[:120]))
        if failed:
            _p("  >> %d gate(s) block this story at stage '%s': %s" % (len(failed), stage, ", ".join(failed)))
            worst = 1
        else:
            _p("  >> stage '%s' fully satisfied" % stage)
    sys.exit(worst if a.strict else 0)

def cmd_build(a):
    dirs = story_dirs(a.root) if a.all else [os.path.join(a.root,"Stories",a.story)]
    for d in dirs: build_story(d)

def cmd_tracker(a):
    wb, rows, n = build_tracker(a.root)
    path = os.path.join(a.root, "Lullable_Story_Card_Tracker.xlsx")
    wb.save(path)
    ready = sum(1 for (m,res,st,rq,fl,fo) in rows if publish_ready(res))
    _p("rebuilt %s from %d manifest(s): %d publish-ready, %d blocked"
       % (os.path.basename(path), n, ready, n-ready))
    for (m,res,st,rq,fl,fo) in rows:
        _p("   %-34s %-12s %s" % (m.get("storyID"), st,
           "READY" if publish_ready(res) else "blocked by " + ",".join(fl) if fl else "stage ok"))

# ================================================================= TRACKER ===
def build_tracker(root):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule

    F, INK, BAND, BAND2, GREY = "Arial", "1F2933", "203D38", "3C5C55", "6B7280"
    RED, GRN, AMB = "FCE4E4", "E3F3E6", "FDF3DC"
    thin = Side(style="thin", color="D5D9D7"); box = Border(thin,thin,thin,thin)
    def font(sz=10,b=False,i=False,c=INK): return Font(name=F,size=sz,bold=b,italic=i,color=c)

    dirs = story_dirs(root)
    rows = []
    for d in dirs:
        m, results = build_story(d, quiet=True)
        stage, required, failed = stage_verdict(m, results)
        rows.append((m, results, stage, required, failed, os.path.basename(d)))

    guidance = {
      "storyID":"never changes after publication","supersedes":"old ID this replaces, if any",
      "workflowStatus":" | ".join(WORKFLOW_STATUSES),"accessDecision":" | ".join(ACCESS_DECISIONS),
      "access":"derived — final only at staging/published","title":"18-48 chars",
      "subtitle":"35-90 chars","narrator":"2-32 chars","genreID_1":"required","genreID_2":"optional",
      "trialPreviewEligible":"true | false","isFeatured":"true | false",
      "publishedAt":"ISO-8601 UTC","colorHex":"6 uppercase hex","accentHex":"6 uppercase hex",
      "durationSeconds":"measured from final audio","bedtimeNote":"90-180 chars","bestFor":"10-24 chars",
      "sleepPace":"8-22 chars","atmosphere":"10-28 chars","description":"220-650 chars",
      "audioAssetID":"immutable once minted","audioMasterFilename":"exact .wav name",
      "commercialRightsStatus":" | ".join(RIGHTS_STATUSES),"audioDelivery":"fixed"}
    widths = {"storyID":26,"supersedes":18,"workflowStatus":14,"accessDecision":14,"access":11,
              "title":30,"subtitle":38,"narrator":16,"genreID_1":16,"genreID_2":16,
              "trialPreviewEligible":18,"isFeatured":11,"publishedAt":21,"colorHex":11,"accentHex":11,
              "durationSeconds":15,"bedtimeNote":44,"bestFor":19,"sleepPace":17,"atmosphere":20,
              "description":48,"audioAssetID":20,"audioMasterFilename":26,
              "commercialRightsStatus":22,"audioDelivery":32}

    wb = Workbook(); ws = wb.active; ws.title = "Story Tracker"
    ws.sheet_view.showGridLines = False
    cols = TRACKER_COLUMNS + ["STAGE","PUBLISH READY"] + [g[0] for g in GATES] + ["manifest"]
    LAST = get_column_letter(len(cols))
    NIN = len(TRACKER_COLUMNS)

    ws.merge_cells("A1:%s1" % LAST)
    c = ws["A1"]; c.value = ("LULLABLE STORY TRACKER  —  GENERATED FROM Stories/<storyID>/story.yaml. "
                             "Do not type here; edit the manifest and run:  lullable.py tracker")
    c.font = Font(name=F,size=12,bold=True,color="FFFFFF")
    c.fill = PatternFill("solid",fgColor=BAND); c.alignment = Alignment("left","center",indent=1)
    ws.row_dimensions[1].height = 26

    groups = [("STORY IDENTITY","A","B"),("WORKFLOW & ACCESS","C","E"),("CARD COPY","F",get_column_letter(NIN)),
              ("VERDICT",get_column_letter(NIN+1),get_column_letter(NIN+2)),
              ("GATES",get_column_letter(NIN+3),LAST)]
    for label,c1,c2 in groups:
        ws.merge_cells("%s2:%s2" % (c1,c2)); cell = ws["%s2" % c1]
        cell.value = label; cell.font = Font(name=F,size=9,bold=True,color="FFFFFF")
        cell.fill = PatternFill("solid",fgColor=BAND2); cell.alignment = Alignment("center","center")
    ws.row_dimensions[2].height = 16

    gate_titles = {g[0]: g[1] for g in GATES}
    for i,name in enumerate(cols):
        L = get_column_letter(i+1)
        ws.column_dimensions[L].width = widths.get(name, 13 if name in gate_titles else 16)
        h = ws["%s3" % L]; h.value = name
        h.font = font(9,b=True); h.alignment = Alignment("left","bottom",wrap_text=True)
        h.fill = PatternFill("solid",fgColor="E7EBE9" if i < NIN else "E2E2E2"); h.border = box
        g = ws["%s4" % L]
        g.value = guidance.get(name, gate_titles.get(name,""))
        g.font = font(8,i=True,c=GREY); g.alignment = Alignment("left","top",wrap_text=True)
        g.fill = PatternFill("solid",fgColor="F5F7F6" if i < NIN else "EFEFEF"); g.border = box
    ws.row_dimensions[3].height = 30; ws.row_dimensions[4].height = 24

    r = 5
    for m, results, stage, required, failed, folder in rows:
        tr = tracker_row(m)
        vals = [tr[c] for c in TRACKER_COLUMNS] + \
               [stage, "READY" if publish_ready(results) else "NOT READY"] + \
               [results[g[0]][1] for g in GATES] + ["Stories/%s/story.yaml" % folder]
        for i,v in enumerate(vals):
            cell = ws.cell(row=r, column=i+1, value=v)
            cell.font = font(9); cell.alignment = Alignment("left","top")
            cell.border = box
            if i >= NIN: cell.alignment = Alignment("center","center")
            if i < NIN: cell.number_format = "0" if TRACKER_COLUMNS[i]=="durationSeconds" else "@"
        ws.row_dimensions[r].height = 26
        r += 1
    last_row = max(r-1, 5)

    # dropdowns WITH error blocking on (this was previously off)
    def dv(formula, colname, prompt):
        if colname not in TRACKER_COLUMNS: return
        L = get_column_letter(TRACKER_COLUMNS.index(colname)+1)
        d = DataValidation(type="list", formula1=formula, allow_blank=False,
                           showDropDown=False, showErrorMessage=True, errorStyle="stop")
        d.error = "Not an allowed value. Pick from the list."; d.errorTitle = "Rejected"
        d.prompt = prompt; d.promptTitle = colname; d.showInputMessage = True
        ws.add_data_validation(d); d.add("%s5:%s%d" % (L,L,last_row+40))
    dv('"%s"' % ",".join(WORKFLOW_STATUSES), "workflowStatus", "Pipeline stage.")
    dv('"%s"' % ",".join(ACCESS_DECISIONS),  "accessDecision", "Free vs premium. Settle before staging.")
    dv('"%s"' % ",".join(GENRES), "genreID_1", "Existing genre IDs only.")
    dv('"%s"' % ",".join(GENRES), "genreID_2", "Existing genre IDs only.")
    dv('"%s"' % ",".join(RIGHTS_STATUSES), "commercialRightsStatus", "Must be verified to publish.")
    dv('"true,false"', "trialPreviewEligible", "true or false.")
    dv('"true,false"', "isFeatured", "true or false.")

    g1 = get_column_letter(NIN+3); rng = "%s5:%s%d" % (g1, LAST, last_row)
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'],
        fill=PatternFill("solid",fgColor=RED), font=Font(name=F,size=9,bold=True,color="B00020")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'],
        fill=PatternFill("solid",fgColor=GRN), font=Font(name=F,size=9,color="1B5E20")))
    pr = get_column_letter(NIN+2); prr = "%s5:%s%d" % (pr,pr,last_row)
    ws.conditional_formatting.add(prr, CellIsRule(operator="equal", formula=['"READY"'],
        fill=PatternFill("solid",fgColor="2E7D32"), font=Font(name=F,size=9,bold=True,color="FFFFFF")))
    ws.conditional_formatting.add(prr, CellIsRule(operator="equal", formula=['"NOT READY"'],
        fill=PatternFill("solid",fgColor=RED), font=Font(name=F,size=9,bold=True,color="B00020")))
    ac = get_column_letter(TRACKER_COLUMNS.index("accessDecision")+1)
    ws.conditional_formatting.add("%s5:%s%d" % (ac,ac,last_row), CellIsRule(operator="equal",
        formula=['"PENDING"'], fill=PatternFill("solid",fgColor=AMB)))

    ws.freeze_panes = "B5"; ws.auto_filter.ref = "A3:%s%d" % (LAST,last_row)

    return wb, rows, len(dirs)

# ============================================================ NEW COMMANDS ===
NARRATION_SKELETON = """Good evening.

PENDING — settling paragraph. Let the day end. Short sentences.

PENDING — second settling paragraph. Hands soft, breathing slow.

PENDING — where we are going tonight, in two or three sentences.

PENDING — permission to drift. "If your attention drifts, let it drift."
So breathe in, slowly, through your nose. And let it go. And let us begin.

§

PENDING — arrival. Establish place, temperature, light, sound.

§

PENDING — descent / moving in closer.

§

PENDING — factual body, section 1. Real mechanics, verified.

§

PENDING — factual body, section 2.

§

PENDING — factual body, section 3.

§

PENDING — factual body, section 4.

§

PENDING — the unresolved or wondering section, if there is one.

§

PENDING — a change of texture: a creature, a person, an object, a sound.

§

PENDING — settle into stillness. Nothing left to watch.

§

PENDING — dissolve. Let it all go, keep nothing.

You are held. You are safe. There is nothing left to do but rest.

Sleep now. Easy, and unhurried, and deep.

Goodnight.
"""

def cmd_new(a):
    sid = story_id_from(a.story_id or a.title)
    d = os.path.join(a.root, "Stories", sid)
    if os.path.exists(d):
        _p("refusing: Stories/%s already exists" % sid); sys.exit(1)
    os.makedirs(os.path.join(d, "audio"))
    m = blank_manifest()
    m["storyID"] = sid
    m["identityResolved"] = not a.unresolved
    m["identityNote"] = a.identity_note or ""
    m["episode"] = {"id": a.id, "pillar": a.pillar, "tags": [t.strip() for t in a.tags.split(",") if t.strip()]}
    m["card"]["title"] = a.title
    if a.genre: m["card"]["genreIDs"] = [a.genre]
    save_manifest(d, m)
    with open(os.path.join(d, "narration.md"), "w", encoding="utf-8") as f:
        f.write(NARRATION_SKELETON)
    with open(os.path.join(d, "audio", "PUT_AUDIO_HERE.txt"), "w", encoding="utf-8") as f:
        f.write("Drop master.wav and delivery.m4a here after the render, then run:\n"
                "  lullable.py closeout %s --voice-id ... --voice-name ... --model eleven_multilingual_v2 \\\n"
                "      --history-id ... --stability 0.45 --similarity 0.75\n" % sid)
    _p("created Stories/%s" % sid)
    _p("  story.yaml     manifest (edit this)")
    _p("  narration.md   skeleton with section seams — replace every PENDING")
    _p("  audio/         master.wav and delivery.m4a go here after the render")
    _p("\nnext:  write the narration, then  lullable.py compile Stories/%s" % sid)

def cmd_status(a):
    dirs = story_dirs(a.root)
    _p("%-34s %-12s %-10s %7s %7s  %s" % ("STORY","STAGE","ACCESS","WORDS","MIN","BLOCKING"))
    _p("-"*104)
    counts = {}
    for d in dirs:
        m = load_manifest(d); results = run_gates(m, d)
        stage, req, failed = stage_verdict(m, results)
        counts[stage] = counts.get(stage, 0) + 1
        ready = publish_ready(results)
        blocking = "READY TO PUBLISH" if ready else (", ".join(failed) if failed else "stage satisfied")
        _p("%-34s %-12s %-10s %7s %7s  %s" % (
            m.get("storyID"), stage, m.get("accessDecision"),
            _get(m,"script.words") or "-", _get(m,"script.estimatedMinutesAt118wpm") or "-", blocking))
    _p("-"*104)
    tot_w = sum((_get(load_manifest(d),"script.words") or 0) for d in dirs)
    tot_m = sum((_get(load_manifest(d),"script.estimatedMinutesAt118wpm") or 0) for d in dirs)
    _p("%d stories · %s words · %.0f minutes · %s" % (
        len(dirs), format(tot_w, ","), tot_m,
        " · ".join("%d %s" % (v,k) for k,v in sorted(counts.items()))))

def cmd_closeout(a):
    d = os.path.join(a.root, "Stories", a.story)
    m = load_manifest(d)
    adir = os.path.join(d, "audio")
    master = a.master or "master.wav"
    delivery = a.delivery or "delivery.m4a"
    mp, dp = os.path.join(adir, master), os.path.join(adir, delivery)
    for label, p in (("master", mp), ("delivery", dp)):
        if not os.path.exists(p):
            _p("cannot close out: %s not found at audio/%s" % (label, os.path.basename(p))); sys.exit(1)
    mpr, dpr = ffprobe(mp), ffprobe(dp)
    if not dpr:
        _p("cannot close out: ffprobe could not read the delivery file"); sys.exit(1)
    m["audio"]["master"] = {"filename": master, "sha256": sha256_of(mp), "bytes": (mpr or {}).get("bytes"),
        "durationSeconds": (mpr or {}).get("durationSeconds"), "codec": (mpr or {}).get("codec"),
        "sampleRate": (mpr or {}).get("sampleRate"), "channels": (mpr or {}).get("channels")}
    m["audio"]["delivery"] = {"filename": delivery, "sha256": sha256_of(dp), "bytes": dpr["bytes"],
        "durationSeconds": dpr["durationSeconds"], "codec": dpr["codec"], "profile": dpr["profile"],
        "sampleRate": dpr["sampleRate"], "channels": dpr["channels"], "bitRateKbps": dpr["bitRateKbps"]}
    m["card"]["durationSeconds"] = int(round(dpr["durationSeconds"]))
    r = m["render"]
    if a.voice_id:    r["voiceId"] = a.voice_id
    if a.voice_name:  r["voiceName"] = a.voice_name
    if a.model:       r["model"] = a.model
    if a.history_id:  r["historyItemId"] = a.history_id
    if a.project_id:  r["projectId"] = a.project_id
    r["renderedAt"] = a.rendered_at or datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    if a.stability  is not None: r["settings"]["stability"] = a.stability
    if a.similarity is not None: r["settings"]["similarityBoost"] = a.similarity
    if a.style      is not None: r["settings"]["style"] = a.style
    r["settings"]["speakerBoost"] = bool(a.speaker_boost)
    if a.narrator: m["card"]["narrator"] = a.narrator
    elif is_placeholder(m["card"]["narrator"]) and a.voice_name: m["card"]["narrator"] = a.voice_name
    if a.asset_id: m["publish"]["audioAssetID"] = a.asset_id
    if m.get("workflowStatus") == "draft": m["workflowStatus"] = "rendered"
    save_manifest(d, m)
    _p("closed out %s" % a.story)
    _p("  duration   %ss  (from the file, not an estimate)" % m["card"]["durationSeconds"])
    _p("  delivery   %s %s %d Hz %d ch %s kbps" % (dpr["codec"], dpr["profile"], dpr["sampleRate"],
                                                   dpr["channels"], dpr["bitRateKbps"]))
    _p("  sha256     master %s…  delivery %s…" % (m["audio"]["master"]["sha256"][:12],
                                                  m["audio"]["delivery"]["sha256"][:12]))
    _p("  status     %s" % m["workflowStatus"])
    build_story(d)
    results = run_gates(m, d); stage, req, failed = stage_verdict(m, results)
    _p("  gates      %s" % ("stage '%s' satisfied" % stage if not failed else
                            "blocked by " + ", ".join(failed)))

def cmd_approve(a):
    d = os.path.join(a.root, "Stories", a.story)
    m = load_manifest(d)
    m["qa"]["audioApproved"] = True
    m["qa"]["approvedBy"] = a.by
    m["qa"]["approvedAt"] = a.at or datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    if a.device:
        m["qa"]["deviceAccepted"] = True
        m["qa"]["deviceNotes"] = a.device_notes or "accepted on a physical device"
    results = run_gates(m, d)
    _, _, failed = stage_verdict(m, results, "qa-approved")
    if not failed and m.get("workflowStatus") in ("draft","rendered"):
        m["workflowStatus"] = "qa-approved"
    save_manifest(d, m); build_story(d)
    _p("QA recorded for %s by %s" % (a.story, a.by))
    _p("  device accepted: %s" % m["qa"]["deviceAccepted"])
    _p("  status: %s%s" % (m["workflowStatus"],
        "" if not failed else "  (still blocked by " + ", ".join(failed) + ")"))

# ========================================================== WEBSITE EXPORT ===
# Generates llulable_website/catalog/<storyID>.md from the manifest.
# The app's four spatial genres are canonical (see Docs/06-decisions.md D19);
# these are their display names on the website.
GENRE_DISPLAY = {"gentle-nature":"Gentle Nature", "ancient-worlds":"Ancient Worlds",
                 "cosmic-journeys":"Cosmic Journeys", "cozy-tales":"Cozy Tales"}
# the website's fixed mood vocabulary, matched from our atmosphere / sleepPace
SITE_MOODS = ("Drifting","Weightless","Wandering","Faraway","Hushed","Dreaming")
MOOD_HINTS = {"Hushed":("hush","quiet","silent","still","soft"),
              "Weightless":("weightless","float","drift down","sink","suspend"),
              "Faraway":("vast","far","cold","distant","immense","deep"),
              "Dreaming":("dream","lamplit","warm","glow","amber"),
              "Wandering":("wander","walk","path","journey","route"),
              "Drifting":("slow","unhurried","gentle","steady","sinking")}

def pick_mood(m):
    blob = " ".join(str(_get(m,"card."+k,"")) for k in ("atmosphere","sleepPace","bestFor")).lower()
    for mood in SITE_MOODS:
        if any(h in blob for h in MOOD_HINTS[mood]): return mood, True
    return "Drifting", False

def _narration_text(story_dir, m):
    p = os.path.join(story_dir, _get(m,"script.narrationFile") or "narration.md")
    if not os.path.exists(p): return ""
    t = open(p, encoding="utf-8").read()
    return re.sub(r'\n\s*§[^\n]*\n', '\n\n', t)

def _paras(t):
    return [p.strip() for p in re.split(r'\n\s*\n', t) if p.strip()]

def first_minute(story_dir, m, words=100):
    """The opening of the narration, after the fixed greeting."""
    ps = _paras(_narration_text(story_dir, m))
    ps = [p for p in ps if not p.startswith("Good evening")]
    out, n = [], 0
    for p in ps:
        out.append(p); n += len(p.split())
        if n >= words: break
    text = " ".join(out).split()
    return " ".join(text[:words]) + ("…" if len(text) > words else "")

def pull_sample(story_dir, m, maxlen=140):
    """A line from the factual body, cut mid-clause, ending on an em-dash."""
    ps = _paras(_narration_text(story_dir, m))
    mid = ps[len(ps)//2:] or ps
    for p in mid:
        for s in split_sentences(p):
            if 14 <= wc(s) <= 40 and not s.startswith(("Let ","So ","And let")):
                s = s.strip()
                if len(s) > maxlen:
                    s = s[:maxlen].rsplit(" ", 1)[0]
                return "…" + s.rstrip(" .,;:") + "—"
    return ""

def ending_giveaway(story_dir, m):
    ps = _paras(_narration_text(story_dir, m))
    for p in reversed(ps):
        if p.startswith(("Sleep now","Goodnight","You are held")): continue
        first = split_sentences(p)[0].strip()
        if wc(first) >= 6: return first
    return ""

def website_markdown(story_dir, m):
    mood, matched = pick_mood(m)
    g = (_get(m,"card.genreIDs") or ["gentle-nature"])[0]
    mins = int(round((_get(m,"card.durationSeconds") or 0) / 60))
    fm = ["---",
          "title: "    + _get(m,"card.title"),
          "narrator: " + _get(m,"card.narrator"),
          "mins: %d" % mins,
          "genre: "    + GENRE_DISPLAY.get(g, g),
          "mood: "     + mood,
          "premium: "  + ("true" if m.get("accessDecision") == "premium" else "false"),
          "date: "     + (_get(m,"card.publishedAt") or "")[:10],
          "blurb: "    + _get(m,"card.bedtimeNote"),
          "sample: "   + pull_sample(story_dir, m),
          "---", ""]
    body = [_get(m,"card.description"), ""]
    end = ending_giveaway(story_dir, m)
    if end:
        body += ["The ending, given away now, per house rules: " + end[0].lower() + end[1:] +
                 " There is nothing else to wait for, which is the point.", ""]
    body += ["## The first minute", "",
             "> " + first_minute(story_dir, m), "",
             "## Why this one works at night", "",
             "%s The pace is %s, and the whole of it is %s. Nothing in it asks to be followed closely, "
             "which is what makes it easy to let go of." % (
                 _get(m,"card.bedtimeNote"),
                 str(_get(m,"card.sleepPace")).lower(),
                 str(_get(m,"card.atmosphere")).lower()), ""]
    return "\n".join(fm + body), (mood, matched)

def cmd_website_export(a):
    dest = os.path.abspath(a.dest)
    dirs = story_dirs(a.root) if a.all else [os.path.join(a.root, "Stories", a.story)]
    wrote, blocked = 0, 0
    for d in dirs:
        m = load_manifest(d); results = run_gates(m, d)
        sid = m.get("storyID")
        # a public page must never carry an estimated duration or an unassigned voice
        reasons = []
        if not isinstance(_get(m,"card.durationSeconds"), (int, float)):
            reasons.append("no measured duration — run closeout after the render")
        if is_placeholder(_get(m,"card.narrator")): reasons.append("narrator not assigned")
        if is_placeholder(_get(m,"card.publishedAt")): reasons.append("no publish date")
        if results["G02"][1] != "PASS": reasons.append("story identity unresolved (G02)")
        if not a.force and reasons:
            _p("  BLOCKED %-34s %s" % (sid, "; ".join(reasons))); blocked += 1; continue
        md, (mood, matched) = website_markdown(d, m)
        if a.dry_run:
            _p("  would write %s/%s.md  (%d bytes)" % (dest, sid, len(md)))
            if a.show: _p("\n" + md + "\n")
        else:
            os.makedirs(dest, exist_ok=True)
            open(os.path.join(dest, sid + ".md"), "w", encoding="utf-8").write(md)
            _p("  wrote %s.md" % sid)
        if not matched: _p("    note: mood defaulted to 'Drifting' — set atmosphere wording to steer it")
        wrote += 1
    _p("\n%d exported, %d blocked" % (wrote, blocked))
    if blocked:
        _p("Blocked stories are correct: the website must not publish an estimated runtime")
        _p("or a narrator nobody has assigned. Render first, then re-run.")

def main():
    ap = argparse.ArgumentParser(description="Lullable pipeline — story.yaml is canonical.")
    ap.add_argument("--root", default=".", help="the lullable_audio folder")
    sub = ap.add_subparsers(dest="cmd", required=True)

    c = sub.add_parser("compile", help="narration -> SSML, updates story.yaml")
    c.add_argument("story_dir"); c.add_argument("--narration", default="narration.md")
    c.add_argument("--seed", type=int, default=11); c.add_argument("--force", action="store_true")
    c.set_defaults(fn=cmd_compile)

    v = sub.add_parser("validate", help="run the staged gate model")
    v.add_argument("story", nargs="?"); v.add_argument("--all", action="store_true")
    v.add_argument("--strict", action="store_true", help="exit non-zero if any required gate fails")
    v.set_defaults(fn=cmd_validate)

    b = sub.add_parser("build", help="regenerate _generated/ from story.yaml")
    b.add_argument("story", nargs="?"); b.add_argument("--all", action="store_true")
    b.set_defaults(fn=cmd_build)

    t = sub.add_parser("tracker", help="rebuild the Excel workbook from all manifests")
    t.set_defaults(fn=cmd_tracker)

    n = sub.add_parser("new", help="scaffold a new episode folder")
    n.add_argument("title"); n.add_argument("--story-id", default="")
    n.add_argument("--genre", default="", choices=("",)+GENRES)
    n.add_argument("--pillar", default=""); n.add_argument("--id", type=int, default=0)
    n.add_argument("--tags", default="sleep story")
    n.add_argument("--unresolved", action="store_true", help="mark identity as unresolved")
    n.add_argument("--identity-note", default="")
    n.set_defaults(fn=cmd_new)

    st = sub.add_parser("status", help="one-screen view of the whole catalogue")
    st.set_defaults(fn=cmd_status)

    co = sub.add_parser("closeout", help="fill the render+audio blocks from the actual files")
    co.add_argument("story")
    co.add_argument("--master", default=""); co.add_argument("--delivery", default="")
    co.add_argument("--voice-id", default=""); co.add_argument("--voice-name", default="")
    co.add_argument("--narrator", default="")
    co.add_argument("--model", default="", choices=("",)+V2_MODELS)
    co.add_argument("--history-id", default=""); co.add_argument("--project-id", default="")
    co.add_argument("--rendered-at", default="")
    co.add_argument("--stability", type=float, default=None)
    co.add_argument("--similarity", type=float, default=None)
    co.add_argument("--style", type=float, default=None)
    co.add_argument("--speaker-boost", action="store_true")
    co.add_argument("--asset-id", default="")
    co.set_defaults(fn=cmd_closeout)

    we = sub.add_parser("website-export", help="generate llulable_website/catalog/*.md from manifests")
    we.add_argument("story", nargs="?"); we.add_argument("--all", action="store_true")
    we.add_argument("--dest", default="../llulable_website/catalog")
    we.add_argument("--dry-run", action="store_true"); we.add_argument("--show", action="store_true")
    we.add_argument("--force", action="store_true", help="export even without measured audio (not advised)")
    we.set_defaults(fn=cmd_website_export)

    ap2 = sub.add_parser("approve", help="record QA and device sign-off")
    ap2.add_argument("story"); ap2.add_argument("--by", required=True)
    ap2.add_argument("--at", default=""); ap2.add_argument("--device", action="store_true")
    ap2.add_argument("--device-notes", default="")
    ap2.set_defaults(fn=cmd_approve)

    a = ap.parse_args(); a.fn(a)

if __name__ == "__main__":
    main()
